import string
from openpyxl.drawing.image import Image
import io
from openpyxl import Workbook


wb = Workbook()
ws = wb.active


def get_pic(url, s):
    r = s.get(url)
    pic = io.BytesIO(r.content)
    return pic


def add_header(videos):
    """add header to excel 

    Args:
        videos = list(json_data["data"]["list"]["vlist"])
    """
    keys = list(videos[0].keys())
    for i, key in enumerate(keys):
        ws.cell(row=1, column=i+1).value = key


def json2excel(videos: list, start_row_index: int, s):
    """
    Args:
        videos = list(json_data["data"]["list"]["vlist"])
        start_row_index (int): _description_
    """
    # video: dict of item
    for video_num, video in enumerate(videos):
        for i, v in enumerate(video.values()):
            row_index = video_num*6+2+start_row_index
            col_index = i+1
            if type(v) == str and v.startswith("http"):
                pic = Image(get_pic(v, s))
                height_origin = pic.height
                pic.height = 100
                pic.width *= 100/height_origin
                ws.add_image(
                    pic, f"{string.ascii_uppercase[col_index]}{row_index}")

            else:
                ws.cell(row=(row_index), column=(col_index), value=v)


def save():
    wb.save("video.xlsx")
